import os
import sys
import io
import pandas as pd
from openpyxl import load_workbook
import re
from datetime import datetime
import logging

# Configuración del logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('process_reports.log'),
        logging.StreamHandler()
    ]
)

# Forzar la codificación UTF-8 en Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

class ReportProcessor:
    def __init__(self, input_dir="xls_folder", temp_dir="reports", output_dir="rapport2"):
        self.input_dir = input_dir
        self.temp_dir = temp_dir
        self.output_dir = output_dir
        self.setup_directories()
        
        self.header_fields = {
            "Report Name": re.compile(r"Report Name\s*:\s*(.*)"),
            "Period": re.compile(r"Period\s*:\s*(.*)"),
            "Domain Name": re.compile(r"Domain Name\s*:\s*(.*)"),
            "Annotation": re.compile(r"Annotation\s*:\s*(.*)"),
            "Number of Records": re.compile(r"Number of Records\s*:\s*(.*)"),
            "Object Name\\(s\\)": re.compile(r"Object Name\\(s\\)\s*:\s*(.*)"),
            "Business Hour Setting": re.compile(r"Business Hour Setting\s*:\s*(.*)"),
            "Filter": re.compile(r"Filter\s*:\s*(.*)"),
            "Generated At": re.compile(r"Generated At\s*:\s*(.*)")
        }
        
        self.message_header = '''\
Cordial Saludo

Debido a los repetidos intentos fallidos de login recibidos, a continuación, les indico los detalles del hallazgo para que por favor procedan a canalizar su solución.

'''
        
        self.message_footer = '''
Después de analizar los logon failures del usuario, se pudo comprobar lo siguiente: Esta cuenta amerita realizar logoff/on de los equipos donde se encuentra logueado actualmente y borrar los datos del credential manager, ya que el mismo puede deberse a cambios recientes en la contraseña o a la necesidad de realizar el cambio.
'''

    def setup_directories(self):
        """Crea los directorios necesarios si no existen."""
        for directory in [self.temp_dir, self.output_dir]:
            os.makedirs(directory, exist_ok=True)

    def extract_header(self, path):
        """Extrae la información del encabezado de un archivo Excel."""
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            lines = [
                str(row[0]) for row in ws.iter_rows(min_row=1, max_row=9, max_col=1, values_only=True)
                if row[0]
            ]
            data = {}
            for line in lines:
                for key, rx in self.header_fields.items():
                    m = rx.search(line)
                    if m:
                        data[key] = m.group(1).strip()
            return data
        except Exception as e:
            logging.error(f"Error al extraer el encabezado de {path}: {e}")
            return {}

    def process_file(self, path):
        """Procesa un archivo de origen (Excel o CSV) para extraer datos."""
        file_ext = os.path.splitext(path)[1].lower()

        if file_ext in ['.xls', '.xlsx']:
            return self._process_excel_file(path)
        elif file_ext == '.csv':
            return self._process_csv_file(path)
        else:
            logging.warning(f"Formato de archivo no soportado: {file_ext}")
            return {}, [], []

    def _process_excel_file(self, path):
        """Procesa un archivo Excel para extraer datos y metadatos."""
        try:
            header = self.extract_header(path)
            
            # Leer la tabla de datos desde la fila 12
            df = pd.read_excel(path, engine="openpyxl", header=None, skiprows=11)
            
            # Extraer IPs (columna B) y razones (columna G)
            ips = []
            reasons = []
            
            if df.shape[1] > 1:
                ips = pd.Series(df.iloc[:, 1].dropna().astype(str)).unique().tolist()
                ips.sort()
            
            if df.shape[1] > 6:
                reasons = pd.Series(df.iloc[:, 6].dropna().astype(str)).unique().tolist()
                reasons.sort()
            
            return header, reasons, ips
        except Exception as e:
            logging.error(f"Error al procesar el archivo Excel {path}: {e}")
            return {}, [], []

    def _process_csv_file(self, path):
        """Procesa un archivo CSV para extraer datos."""
        try:
            header = {}  # No hay metadata de cabecera en CSV
            
            # Asumimos que la primera fila es el encabezado de la tabla
            # y que las IPs están en la columna 2 (índice 1) y las razones en la 7 (índice 6)
            df = pd.read_csv(path, header=0, on_bad_lines='skip', encoding='utf-8')
            
            ips = []
            reasons = []
            
            # Las columnas se leen por nombre si existen, si no por indice.
            if 'Client IP' in df.columns and 'Reason' in df.columns:
                 ips = pd.Series(df['Client IP'].dropna().astype(str)).unique().tolist()
                 ips.sort()
                 reasons = pd.Series(df['Reason'].dropna().astype(str)).unique().tolist()
                 reasons.sort()
            else: # Fallback a indices si las columnas no tienen nombre
                if df.shape[1] > 1:
                    ips = pd.Series(df.iloc[:, 1].dropna().astype(str)).unique().tolist()
                    ips.sort()
                
                if df.shape[1] > 6:
                    reasons = pd.Series(df.iloc[:, 6].dropna().astype(str)).unique().tolist()
                    reasons.sort()

            return header, reasons, ips
        except Exception as e:
            logging.error(f"Error al procesar el archivo CSV {path}: {e}")
            return {}, [], []

    def generate_intermediate_report(self, filename, header, reasons, ips):
        """Genera un informe intermedio en formato de texto."""
        base = os.path.basename(filename)
        name, _ = os.path.splitext(base)
        out_path = os.path.join(self.temp_dir, f"{name}_reporte.txt")
        
        try:
            with open(out_path, "w", encoding="utf-8") as f:
                f.write("="*50 + "\n\n")
                f.write("=== Encabezado ===\n")
                for k in self.header_fields:
                    f.write(f"{k}: {header.get(k, '<no encontrado>')}\n")
                f.write("\n=== Razones de Fallo Únicas ===\n")
                for fr in reasons:
                    f.write(f"- {fr}\n")
                f.write("\n=== IPs de Clientes Únicas ===\n")
                for ip in ips:
                    f.write(f"- {ip}\n")
            
            logging.info(f"Reporte intermedio generado: {out_path}")
            return out_path
        except Exception as e:
            logging.error(f"Error al generar el reporte para {filename}: {e}")
            return None

    def generate_final_report(self, intermediate_path):
        """Genera el informe final combinando el contenido intermedio con mensajes predefinidos."""
        try:
            with open(intermediate_path, "r", encoding="utf-8") as f:
                report_content = f.read()
            
            combined = self.message_header + report_content + "\n" + self.message_footer
            
            base = os.path.splitext(os.path.basename(intermediate_path))[0]
            out_fname = f"{base}_final.txt"
            out_path = os.path.join(self.output_dir, out_fname)
            
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(combined)
            
            logging.info(f"Reporte final generado: {out_path}")
            return out_path
        except Exception as e:
            logging.error(f"Error al generar el reporte final: {e}")
            return None

    def extract_intermediate_reports(self):
        """
        Procesa todos los archivos de origen y genera solo los reportes intermedios.
        Devuelve listas de archivos procesados y fallidos.
        """
        processed_files = []
        failed_files = []
        
        if not os.path.exists(self.input_dir):
            logging.error(f"El directorio de entrada {self.input_dir} no existe.")
            return processed_files, failed_files
        
        source_files = [f for f in os.listdir(self.input_dir) if f.lower().endswith(('.xls', '.xlsx', '.csv'))]
        
        if not source_files:
            logging.warning(f"No se encontraron archivos Excel o CSV en {self.input_dir}")
            return processed_files, failed_files
        
        logging.info(f"Extrayendo información de {len(source_files)} archivos...")
        
        for filename in source_files:
            file_path = os.path.join(self.input_dir, filename)
            logging.info(f"Procesando: {filename}")
            
            try:
                header, reasons, ips = self.process_file(file_path)
                intermediate_path = self.generate_intermediate_report(file_path, header, reasons, ips)
                
                if intermediate_path:
                    processed_files.append(filename)
                    logging.info(f"✓ Información extraída de {filename}")
                else:
                    failed_files.append(filename)
                    
            except Exception as e:
                logging.error(f"Error al procesar {filename}: {e}")
                failed_files.append(filename)
        
        return processed_files, failed_files

    def generate_final_reports(self):
        """
        Genera los reportes finales a partir de los reportes intermedios existentes.
        """
        processed_final = []
        failed_final = []
        
        if not os.path.exists(self.temp_dir):
            logging.warning(f"El directorio temporal {self.temp_dir} no existe. No hay nada que procesar.")
            return processed_final, failed_final
            
        intermediate_files = [f for f in os.listdir(self.temp_dir) if f.endswith("_reporte.txt")]
        
        if not intermediate_files:
            logging.warning("No se encontraron reportes intermedios para procesar.")
            return processed_final, failed_final
            
        logging.info(f"Generando {len(intermediate_files)} reportes finales...")
        
        for filename in intermediate_files:
            intermediate_path = os.path.join(self.temp_dir, filename)
            try:
                final_path = self.generate_final_report(intermediate_path)
                if final_path:
                    processed_final.append(final_path)
                else:
                    failed_final.append(filename)
            except Exception as e:
                logging.error(f"Error al generar el reporte final para {filename}: {e}")
                failed_final.append(filename)
                
        # Limpiar los archivos intermedios después de usarlos
        self.cleanup_temp_files()
        
        return processed_final, failed_final

    def cleanup_temp_files(self):
        """Limpia los archivos temporales generados."""
        try:
            for filename in os.listdir(self.temp_dir):
                if filename.endswith("_reporte.txt"):
                    os.remove(os.path.join(self.temp_dir, filename))
            logging.info("Archivos temporales limpiados.")
        except Exception as e:
            logging.error(f"Error durante la limpieza de archivos temporales: {e}")

    def run(self, cleanup=True):
        """Ejecuta el proceso completo de generación de reportes (flujo original)."""
        start_time = datetime.now()
        logging.info("=== Inicio del procesamiento de reportes ===")
        
        processed_intermediate, failed_intermediate = self.extract_intermediate_reports()
        
        if processed_intermediate:
            processed_final, failed_final = self.generate_final_reports()
        else:
            processed_final, failed_final = [], []

        end_time = datetime.now()
        duration = end_time - start_time
        
        logging.info("=== Resumen del procesamiento ===")
        # Combina las listas para el resumen
        processed_count = len(processed_final)
        failed_count = len(failed_intermediate) + len(failed_final)
        failed_files_list = failed_intermediate + failed_final
        
        logging.info(f"Archivos procesados con éxito: {processed_count}")
        logging.info(f"Archivos con errores: {failed_count}")
        logging.info(f"Duración total: {duration}")
        
        if failed_files_list:
            logging.warning(f"Archivos con errores: {', '.join(set(failed_files_list))}")
        
        return {
            'processed': processed_final,
            'failed': failed_files_list,
            'duration': duration
        }

def main():
    """Función principal para ejecución directa."""
    processor = ReportProcessor()
    results = processor.run()
    
    print(f"\n🎯 ¡Procesamiento completado!")
    print(f"📊 {len(results['processed'])} archivos procesados")
    print(f"❌ {len(results['failed'])} archivos con errores")
    print(f"⏱️ Duración: {results['duration']}")

if __name__ == "__main__":
    main()